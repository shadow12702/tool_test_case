from __future__ import annotations

import csv
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from model.types import PromptItem
from service.file_utils import looks_like_xlsx


PROMPT_COLUMN_CANDIDATES = ("Prompt", "user_input", "Prompt (mới)", "Prompt (Cũ)")
PROMPT_ID_COLUMN_CANDIDATES = (
    "Prompt_ID",
    "Prompt-ID",
    "Prompt ID",
    "PromptID",
    "Type ID",
    "TypeID",
    "type_id",
    "Type ID New",
)


def _read_prompts_from_xlsx_bytes(
    data: bytes, source_name: str, only_sheet: str | None = None,
) -> list[PromptItem]:
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    prompts: list[PromptItem] = []

    sheets_to_read = [only_sheet] if only_sheet and only_sheet in wb.sheetnames else wb.sheetnames
    for sheet_name in sheets_to_read:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            continue

        header_norm = [("" if v is None else str(v).strip()) for v in header]
        col_idx = None
        for cand in PROMPT_COLUMN_CANDIDATES:
            if cand in header_norm:
                col_idx = header_norm.index(cand)
                break
        if col_idx is None:
            continue

        prompt_id_idx = None
        for cand in PROMPT_ID_COLUMN_CANDIDATES:
            if cand in header_norm:
                prompt_id_idx = header_norm.index(cand)
                break

        row_index = 1
        for row in rows:
            row_index += 1
            if not row or col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None:
                continue
            text = str(val).strip()
            if not text:
                continue
            prompt_id = ""
            if prompt_id_idx is not None and prompt_id_idx < len(row) and row[prompt_id_idx] is not None:
                prompt_id = str(row[prompt_id_idx]).strip()
            prompts.append(
                PromptItem(
                    sheet=f"{source_name}:{sheet_name}",
                    row_index=row_index,
                    user_input=text,
                    prompt_id=prompt_id,
                )
            )

    return prompts


def _read_prompts_from_csv_text(text: str, source_name: str) -> list[PromptItem]:
    # Try DictReader first (expects header)
    lines = [ln for ln in text.splitlines() if ln.strip()]
    if not lines:
        return []

    reader = csv.DictReader(lines)
    if reader.fieldnames:
        fields = [f.strip() for f in reader.fieldnames if f]
        chosen = None
        for cand in PROMPT_COLUMN_CANDIDATES:
            if cand in fields:
                chosen = cand
                break
        prompt_id_field = None
        for cand in PROMPT_ID_COLUMN_CANDIDATES:
            if cand in fields:
                prompt_id_field = cand
                break
        if chosen:
            out: list[PromptItem] = []
            for i, row in enumerate(reader, start=2):  # header is row 1
                val = (row.get(chosen) or "").strip()
                if val:
                    prompt_id = (
                        (row.get(prompt_id_field) or "").strip()
                        if prompt_id_field
                        else ""
                    )
                    out.append(
                        PromptItem(
                            sheet=f"{source_name}:CSV",
                            row_index=i,
                            user_input=val,
                            prompt_id=prompt_id,
                        )
                    )
            return out

    # Fallback: single-column CSV (no header)
    out2: list[PromptItem] = []
    for i, ln in enumerate(lines, start=1):
        out2.append(
            PromptItem(
                sheet=f"{source_name}:CSV",
                row_index=i,
                user_input=ln.strip(),
                prompt_id="",
            )
        )
    return [p for p in out2 if p.user_input]


def read_prompts_from_file(
    path: Path, sheet_name: str | None = None,
) -> list[PromptItem]:
    """
    Reads prompts from:
      - real XLSX
      - real CSV (UTF-8)
      - mislabeled XLSX saved with .csv extension (detected by ZIP magic)

    If sheet_name is provided, only that sheet is read from the workbook.
    Column `Prompt` is `user_input`.
    """
    source_name = path.name

    if path.suffix.lower() in (".xlsx", ".xlsm", ".xltx", ".xltm") or looks_like_xlsx(path):
        data = path.read_bytes()
        return _read_prompts_from_xlsx_bytes(
            data, source_name=source_name, only_sheet=sheet_name,
        )

    # Treat as text CSV (sheet_name ignored)
    text = path.read_text(encoding="utf-8", errors="replace")
    return _read_prompts_from_csv_text(text, source_name=source_name)

