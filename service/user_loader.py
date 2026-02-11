from __future__ import annotations

import csv
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from model.types import UserItem
from service.file_utils import looks_like_xlsx


USER_ID_CANDIDATES = ("user-id", "user_id", "userid", "userId")
USER_NAME_CANDIDATES = ("user-name", "user_name", "username", "userName")


def _read_users_from_xlsx_bytes(data: bytes) -> list[UserItem]:
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []

    header_norm = [("" if v is None else str(v).strip()) for v in header]

    def find_idx(cands: tuple[str, ...]) -> int | None:
        for c in cands:
            if c in header_norm:
                return header_norm.index(c)
        return None

    id_idx = find_idx(USER_ID_CANDIDATES)
    name_idx = find_idx(USER_NAME_CANDIDATES)
    if id_idx is None or name_idx is None:
        return []

    out: list[UserItem] = []
    for row in rows:
        if not row:
            continue
        user_id = "" if id_idx >= len(row) or row[id_idx] is None else str(row[id_idx]).strip()
        user_name = "" if name_idx >= len(row) or row[name_idx] is None else str(row[name_idx]).strip()
        if user_id and user_name:
            out.append(UserItem(user_id=user_id, user_name=user_name))
    return out


def _read_users_from_csv_text(text: str) -> list[UserItem]:
    lines = [ln for ln in text.splitlines() if ln.strip()]
    if not lines:
        return []

    reader = csv.DictReader(lines)
    fields = [f.strip() for f in (reader.fieldnames or []) if f]
    if fields:
        def pick_field(cands: tuple[str, ...]) -> str | None:
            for c in cands:
                if c in fields:
                    return c
            return None

        id_field = pick_field(USER_ID_CANDIDATES)
        name_field = pick_field(USER_NAME_CANDIDATES)
        if id_field and name_field:
            out: list[UserItem] = []
            for row in reader:
                user_id = (row.get(id_field) or "").strip()
                user_name = (row.get(name_field) or "").strip()
                if user_id and user_name:
                    out.append(UserItem(user_id=user_id, user_name=user_name))
            if out:
                return out

    # Fallback: manual entry without header, each line like:
    #   userid-001,user-001
    # or separated by tab/semicolon/pipe.
    out2: list[UserItem] = []
    for ln in lines:
        # normalize common separators
        for sep in ("\t", ";", "|"):
            ln = ln.replace(sep, ",")
        parts = [p.strip() for p in ln.split(",") if p.strip()]
        if len(parts) < 2:
            continue
        user_id, user_name = parts[0], parts[1]
        # Skip obvious header-like lines
        if user_id.lower() in USER_ID_CANDIDATES and user_name.lower() in USER_NAME_CANDIDATES:
            continue
        out2.append(UserItem(user_id=user_id, user_name=user_name))
    return out2


def load_users_from_list(path: Path) -> list[UserItem]:
    """
    Read users from either:
    - XLSX
    - CSV (utf-8)
    - XLSX but named `.csv` (detected by ZIP magic)

    Expected columns:
    - user-id / user_name variants
    """
    if not path.exists():
        return []

    if path.suffix.lower() in (".xlsx", ".xlsm", ".xltx", ".xltm") or looks_like_xlsx(path):
        return _read_users_from_xlsx_bytes(path.read_bytes())

    text = path.read_text(encoding="utf-8", errors="replace")
    return _read_users_from_csv_text(text)

